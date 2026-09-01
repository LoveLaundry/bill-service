from datetime import datetime, timezone
from typing import Dict, List, Optional
import csv
import io
from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from ..auth_helper import get_current_user, require_capability
from ..crypto_helper import decrypt_dict, get_search_token
from ..database.main_db import (
    audit_collection,
    bills_collection,
    deliveries_collection,
    gatepasses_collection,
    payments_collection,
)

router = APIRouter(tags=["dashboard"])

SENSITIVE_FIELDS_GP = ["client_name", "items", "notes"]
SENSITIVE_FIELDS_DEL = ["client_name", "items", "notes"]
SENSITIVE_FIELDS_BILL = ["client_name", "quotation_title", "notes", "items"]
SENSITIVE_FIELDS_PAY = ["client_name", "notes"]


def _decrypt_gp(doc: dict) -> dict:
    dec = decrypt_dict(doc, SENSITIVE_FIELDS_GP)
    dec["id"] = str(dec["_id"])
    del dec["_id"]
    return dec


def _decrypt_del(doc: dict) -> dict:
    dec = decrypt_dict(doc, SENSITIVE_FIELDS_DEL)
    dec["id"] = str(dec["_id"])
    del dec["_id"]
    return dec


def _decrypt_bill(doc: dict) -> dict:
    dec = decrypt_dict(doc, SENSITIVE_FIELDS_BILL)
    dec["id"] = str(dec["_id"])
    del dec["_id"]
    return dec


def _decrypt_pay(doc: dict) -> dict:
    dec = decrypt_dict(doc, SENSITIVE_FIELDS_PAY)
    dec["id"] = str(dec["_id"])
    del dec["_id"]
    return dec


# --- 1. Client Dashboard Summary ---
@router.get(
    "/dashboard/client-summary",
    dependencies=[Depends(require_capability("dashboard:read"))],
)
async def get_client_summary(client_name: str = Query(...)):
    token = get_search_token(client_name)

    gp_cursor = gatepasses_collection.find({"client_name_search": token})
    del_cursor = deliveries_collection.find({"client_name_search": token})
    bill_cursor = bills_collection.find({"client_name_search": token})
    pay_cursor = payments_collection.find({"client_name_search": token})

    gps = []
    async for d in gp_cursor:
        try:
            gps.append(_decrypt_gp(d))
        except Exception:
            pass

    deliveries = []
    async for d in del_cursor:
        try:
            deliveries.append(_decrypt_del(d))
        except Exception:
            pass

    bills = []
    async for d in bill_cursor:
        try:
            bills.append(_decrypt_bill(d))
        except Exception:
            pass

    payments = []
    async for d in pay_cursor:
        try:
            payments.append(_decrypt_pay(d))
        except Exception:
            pass

    # Compute stats
    total_received = 0
    open_mismatches = 0
    mismatches_list = []
    balances_map: Dict[str, dict] = {}

    for gp in gps:
        for item in gp.get("items", []):
            name = item["item_name"]
            client_qty = item.get("client_qty", 0)
            rec_qty = item.get("received_qty", 0)
            diff = item.get("difference", 0)
            total_received += rec_qty

            if diff != 0:
                open_mismatches += 1
                mismatches_list.append(
                    {
                        "gate_pass_id": gp["id"],
                        "gate_pass_number": gp["gate_pass_number"],
                        "item_name": name,
                        "expected": client_qty,
                        "received": rec_qty,
                        "difference": diff,
                        "reason": item.get("mismatch_reason", "OTHER"),
                        "notes": item.get("mismatch_notes"),
                        "date": gp.get("receiving_date"),
                    }
                )

            if name not in balances_map:
                balances_map[name] = {"received": 0, "delivered": 0, "pending": 0}
            balances_map[name]["received"] += rec_qty

    total_delivered = 0
    for dl in deliveries:
        if dl.get("status") == "CANCELLED":
            continue
        for item in dl.get("items", []):
            name = item["item_name"]
            qty = item["quantity"]
            total_delivered += qty
            if name not in balances_map:
                balances_map[name] = {"received": 0, "delivered": 0, "pending": 0}
            balances_map[name]["delivered"] += qty

    pending_items_sum = 0
    for name, item_bal in balances_map.items():
        item_bal["pending"] = max(0, item_bal["received"] - item_bal["delivered"])
        pending_items_sum += item_bal["pending"]

    total_billed = 0.0
    total_paid = 0.0
    outstanding_amt = 0.0
    pending_bills_count = 0
    for bill in bills:
        if bill.get("payment_status") == "CANCELLED":
            continue
        total_billed += bill.get("grand_total", 0.0)
        total_paid += bill.get("paid_amount", 0.0)
        outstanding_amt += bill.get("outstanding_amount", 0.0)
        if bill.get("payment_status") not in ["PAID"]:
            pending_bills_count += 1

    return {
        "stats": {
            "total_gate_passes": len(gps),
            "total_items_received": total_received,
            "total_items_delivered": total_delivered,
            "pending_items": pending_items_sum,
            "open_mismatches": open_mismatches,
            "pending_bills": pending_bills_count,
            "total_billed": round(total_billed, 2),
            "total_paid": round(total_paid, 2),
            "outstanding_amount": round(outstanding_amt, 2),
        },
        "gatepasses": gps,
        "recent_gate_passes": gps,  # Include both for compatibility
        "deliveries": deliveries,
        "mismatches": mismatches_list,
        "pending_balances": [
            {"item_name": k, **v} for k, v in balances_map.items()
        ],
        "bills": bills,
        "recent_bills": bills,  # Include both for compatibility
        "payments": payments,
    }


# --- 2. Client-wise Report ---
@router.get(
    "/reports/client-wise",
    dependencies=[Depends(require_capability("dashboard:read"))],
)
async def get_client_wise_report():
    gp_cursor = gatepasses_collection.find()
    clients_map: Dict[str, dict] = {}

    async for doc in gp_cursor:
        try:
            gp = _decrypt_gp(doc)
            c_label = gp["client_name"].strip()
            if not c_label:
                continue
            if c_label not in clients_map:
                clients_map[c_label] = {
                    "client_name": c_label,
                    "total_received": 0,
                    "total_delivered": 0,
                    "total_pending": 0,
                    "total_mismatches": 0,
                    "total_billed": 0.0,
                    "paid_amount": 0.0,
                    "outstanding": 0.0,
                    "gate_pass_count": 0,
                    "items_detail": {},
                    "gate_passes": [],
                }
            clients_map[c_label]["gate_pass_count"] += 1
            gp_items = []
            for item in gp.get("items", []):
                clients_map[c_label]["total_received"] += item.get("received_qty", 0)
                if item.get("difference", 0) != 0:
                    clients_map[c_label]["total_mismatches"] += 1
                item_key = item.get("item_name", "")
                spec = item.get("specification") or ""
                detail_key = f"{item_key}||{spec}" if spec else item_key
                if detail_key not in clients_map[c_label]["items_detail"]:
                    clients_map[c_label]["items_detail"][detail_key] = {
                        "item_name": item_key,
                        "specification": spec,
                        "category": item.get("category") or "",
                        "received": 0,
                        "delivered": 0,
                    }
                clients_map[c_label]["items_detail"][detail_key]["received"] += item.get("received_qty", 0)
                gp_items.append({
                    "item_name": item_key,
                    "specification": spec,
                    "received": item.get("received_qty", 0),
                })
            clients_map[c_label]["gate_passes"].append({
                "gate_pass_number": gp.get("gate_pass_number", ""),
                "receiving_date": str(gp.get("receiving_date", ""))[:10],
                "items": gp_items,
            })
        except Exception:
            continue

    del_cursor = deliveries_collection.find({"status": {"$ne": "CANCELLED"}})
    async for doc in del_cursor:
        try:
            dl = _decrypt_del(doc)
            client = dl["client_name"].strip()
            if client in clients_map:
                for item in dl.get("items", []):
                    clients_map[client]["total_delivered"] += item.get("quantity", 0)
                    item_key = item.get("item_name", "")
                    spec = item.get("specification") or ""
                    detail_key = f"{item_key}||{spec}" if spec else item_key
                    if detail_key in clients_map[client]["items_detail"]:
                        clients_map[client]["items_detail"][detail_key]["delivered"] += item.get("quantity", 0)
        except Exception:
            continue

    bill_cursor = bills_collection.find({"payment_status": {"$ne": "CANCELLED"}})
    async for doc in bill_cursor:
        try:
            bill = _decrypt_bill(doc)
            client = bill["client_name"].strip()
            if client in clients_map:
                clients_map[client]["total_billed"] += bill.get("grand_total", 0.0)
                clients_map[client]["paid_amount"] += bill.get("paid_amount", 0.0)
                clients_map[client]["outstanding"] += bill.get("outstanding_amount", 0.0)
        except Exception:
            continue

    results = []
    for c_label, stats in clients_map.items():
        stats["total_pending"] = max(0, stats["total_received"] - stats["total_delivered"])
        stats["total_billed"] = round(stats["total_billed"], 2)
        stats["paid_amount"] = round(stats["paid_amount"], 2)
        stats["outstanding"] = round(stats["outstanding"], 2)
        items_list = list(stats.pop("items_detail").values())
        for it in items_list:
            it["pending"] = max(0, it["received"] - it["delivered"])
        stats["items"] = [it for it in items_list if it["pending"] > 0]
        stats["gate_passes"] = stats.pop("gate_passes", [])
        results.append(stats)

    return results


# --- 3. Item-wise Report ---
@router.get(
    "/reports/item-wise",
    dependencies=[Depends(require_capability("dashboard:read"))],
)
async def get_item_wise_report():
    gp_cursor = gatepasses_collection.find()
    items_map: Dict[str, dict] = {}

    async for doc in gp_cursor:
        try:
            gp = _decrypt_gp(doc)
            for item in gp.get("items", []):
                name = item["item_name"]
                if name not in items_map:
                    items_map[name] = {
                        "item_name": name,
                        "total_received": 0,
                        "total_delivered": 0,
                        "pending": 0,
                        "mismatch_count": 0,
                        "clients": set()
                    }
                items_map[name]["total_received"] += item.get("received_qty", 0)
                if item.get("difference", 0) != 0:
                    items_map[name]["mismatch_count"] += 1
                items_map[name]["clients"].add(gp.get("client_name", ""))
        except Exception:
            pass

    del_cursor = deliveries_collection.find({"status": {"$ne": "CANCELLED"}})
    async for doc in del_cursor:
        try:
            dl = _decrypt_del(doc)
            for item in dl.get("items", []):
                name = item["item_name"]
                if name in items_map:
                    items_map[name]["total_delivered"] += item.get("quantity", 0)
        except Exception:
            pass

    results = []
    for name, stats in items_map.items():
        stats["pending"] = max(0, stats["total_received"] - stats["total_delivered"])
        stats["client_count"] = len(stats["clients"])
        del stats["clients"]  # Remove the set before returning
        results.append(stats)

    return results


# --- 4. Gate Pass-wise Report ---
@router.get(
    "/reports/gatepass-wise",
    dependencies=[Depends(require_capability("dashboard:read"))],
)
async def get_gatepass_wise_report():
    gp_cursor = gatepasses_collection.find().sort("receiving_date", -1)
    results = []

    async for doc in gp_cursor:
        try:
            gp = _decrypt_gp(doc)
            gp_id = gp["id"]

            total_received = sum(x.get("received_qty", 0) for x in gp.get("items", []))
            mismatch_count = sum(1 for x in gp.get("items", []) if x.get("difference", 0) != 0)

            del_ids = []
            del_cursor_2 = deliveries_collection.find(
                {"gate_pass_id": gp_id, "status": {"$ne": "CANCELLED"}}
            )
            total_delivered = 0
            async for d_doc in del_cursor_2:
                try:
                    dl = _decrypt_del(d_doc)
                    del_ids.append(dl["id"])
                    total_delivered += sum(x.get("quantity", 0) for x in dl.get("items", []))
                except Exception:
                    pass

            results.append(
                {
                    "gate_pass_number": gp["gate_pass_number"],
                    "client_name": gp["client_name"],
                    "receiving_date": gp.get("receiving_date"),
                    "received_by": gp.get("received_by"),
                    "total_received": total_received,
                    "total_delivered": total_delivered,
                    "mismatch_count": mismatch_count,
                    "status": gp.get("status"),
                }
            )
        except Exception:
            pass

    return results


# --- 5. Billing Summary Report ---
@router.get(
    "/reports/billing",
    dependencies=[Depends(require_capability("dashboard:read"))],
)
async def get_billing_report():
    cursor = bills_collection.find({"payment_status": {"$ne": "CANCELLED"}})
    total_sales = 0.0
    pending_amount = 0.0
    paid_amount = 0.0
    outstanding_amt = 0.0

    async for doc in cursor:
        try:
            bill = _decrypt_bill(doc)
            total_sales += bill.get("grand_total", 0.0)
            paid_amount += bill.get("paid_amount", 0.0)
            outstanding_amt += bill.get("outstanding_amount", 0.0)
            if bill.get("payment_status") != "PAID":
                pending_amount += bill.get("outstanding_amount", 0.0)
        except Exception:
            pass

    return {
        "total_sales": round(total_sales, 2),
        "pending_bills_amount": round(pending_amount, 2),
        "paid_bills_amount": round(paid_amount, 2),
        "outstanding_amount": round(outstanding_amt, 2),
    }


# --- 6. Audit Logs ---
@router.get(
    "/audit-logs",
    dependencies=[Depends(require_capability("dashboard:read"))],
)
async def get_audit_logs(
    limit: int = Query(50, ge=1, le=200),
):
    cursor = audit_collection.find().sort("timestamp", -1).limit(limit)
    logs = []
    async for doc in cursor:
        doc["id"] = str(doc["_id"])
        del doc["_id"]
        logs.append(doc)
    return logs

from datetime import timedelta

BILL_DECRYPT_FIELDS = ["client_name"]
GP_DECRYPT_FIELDS = ["client_name", "items"]
DELIVERY_DECRYPT_FIELDS = ["items"]

PERIOD_DAYS = {
    "day": 1,
    "week": 7,
    "month": 30,
    "quarter": 90,
    "year": 365,
}


def _utc(dt: Optional[datetime]) -> Optional[datetime]:
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _windows(period: str):
    now = datetime.now(timezone.utc)
    span = PERIOD_DAYS.get(period, 30)
    cur_start = now - timedelta(days=span)
    prev_end = cur_start - timedelta(microseconds=1)
    prev_start = prev_end - timedelta(days=span)
    return cur_start, now, prev_start, prev_end


async def _fetch_bills(start: datetime, end: datetime):
    docs = []
    cursor = bills_collection.find(
        {"created_at": {"$gte": start, "$lte": end}},
        {
            "client_name": 1,
            "grand_total": 1,
            "paid_amount": 1,
            "outstanding_amount": 1,
            "payment_status": 1,
            "created_at": 1,
            "encryption_metadata": 1,
            "_id": 1,
        },
    )
    async for d in cursor:
        try:
            dec = decrypt_dict(d, BILL_DECRYPT_FIELDS)
            client = dec.get("client_name") or "Unknown"
        except Exception:
            client = "Unknown"
        docs.append(
            {
                "client_name": client,
                "grand_total": float(d.get("grand_total") or 0),
                "paid_amount": float(d.get("paid_amount") or 0),
                "outstanding_amount": float(d.get("outstanding_amount") or 0),
                "payment_status": (d.get("payment_status") or "DRAFT"),
                "created_at": _utc(d.get("created_at")),
            }
        )
    return docs


async def _fetch_gate_passes(start: datetime, end: datetime):
    out = []
    cursor = gatepasses_collection.find(
        {"created_at": {"$gte": start, "$lte": end}},
        {"client_name": 1, "items": 1, "gate_pass_number": 1, "created_at": 1, "encryption_metadata": 1, "_id": 1},
    )
    async for d in cursor:
        try:
            dec = decrypt_dict(d, GP_DECRYPT_FIELDS)
            client = dec.get("client_name") or "Unknown"
            items = dec.get("items") or []
        except Exception:
            client = "Unknown"
            items = []
        out.append(
            {
                "id": str(d.get("_id")),
                "gate_pass_number": d.get("gate_pass_number"),
                "client_name": client,
                "created_at": _utc(d.get("created_at")),
                "items": [
                    {
                        "item_name": it.get("item_name"),
                        "received_qty": int(it.get("received_qty") or 0),
                    }
                    for it in items
                ],
            }
        )
    return out


async def _fetch_deliveries(gate_pass_ids: List[str]):
    if not gate_pass_ids:
        return []
    out = []
    cursor = deliveries_collection.find(
        {"gate_pass_id": {"$in": gate_pass_ids}},
        {"gate_pass_id": 1, "items": 1, "created_at": 1, "encryption_metadata": 1, "_id": 1},
    )
    async for d in cursor:
        try:
            dec = decrypt_dict(d, DELIVERY_DECRYPT_FIELDS)
            items = dec.get("items") or []
        except Exception:
            items = []
        out.append(
            {
                "gate_pass_id": d.get("gate_pass_id"),
                "items": [
                    {
                        "item_name": it.get("item_name"),
                        "quantity": int(it.get("quantity") or 0),
                    }
                    for it in items
                ],
            }
        )
    return out


def _bucket_key(dt: Optional[datetime], period: str):
    if dt is None:
        return None
    if period == "day":
        h = f"{dt.hour:02d}:00"
        return (h, h)
    if period in ("quarter", "year"):
        key = f"{dt.year}-{dt.month:02d}"
        return (key, key)
    key = dt.date().isoformat()
    return (key, f"{dt.month}/{dt.day}")


def _outstanding(b):
    if b["outstanding_amount"]:
        return b["outstanding_amount"]
    return max(0.0, b["grand_total"] - b["paid_amount"])


def aggregate(bills, gate_passes, deliveries, period):
    revenue = sum(b["grand_total"] for b in bills)
    collected = sum(b["paid_amount"] for b in bills)
    outstanding = sum(_outstanding(b) for b in bills)
    bill_count = len(bills)
    paid = sum(1 for b in bills if b["payment_status"] == "PAID")
    partial = sum(1 for b in bills if b["payment_status"] == "PARTIALLY_PAID")
    pending = sum(
        1
        for b in bills
        if b["payment_status"] not in ("PAID", "PARTIALLY_PAID", "CANCELLED")
    )
    collection_rate = (collected / revenue * 100) if revenue else 0.0
    avg = (revenue / bill_count) if bill_count else 0.0

    gp_count = len(gate_passes)
    items_received = sum(it["received_qty"] for gp in gate_passes for it in gp["items"])

    del_map = {}
    for d in deliveries:
        m = del_map.setdefault(d["gate_pass_id"], {})
        for it in d["items"]:
            m[it["item_name"]] = m.get(it["item_name"], 0) + it["quantity"]
    items_delivered = sum(it["quantity"] for d in deliveries for it in d["items"])
    items_pending = max(0, items_received - items_delivered)

    active_clients = len(set(b["client_name"] for b in bills))

    client_agg = {}
    for b in bills:
        c = client_agg.setdefault(
            b["client_name"], {"revenue": 0.0, "outstanding": 0.0, "bills": 0}
        )
        c["revenue"] += b["grand_total"]
        c["outstanding"] += _outstanding(b)
        c["bills"] += 1
    top_clients = sorted(
        ({"client_name": k, **v} for k, v in client_agg.items()),
        key=lambda x: -x["revenue"],
    )[:6]

    series = {}
    for b in bills:
        k = _bucket_key(b["created_at"], period)
        if not k:
            continue
        key, label = k
        e = series.setdefault(key, {"label": label, "revenue": 0.0, "collected": 0.0})
        e["revenue"] += b["grand_total"]
        e["collected"] += b["paid_amount"]
    revenue_series = sorted(
        ({"key": k, **v} for k, v in series.items()), key=lambda x: x["key"]
    )

    st = {"PAID": 0.0, "PARTIAL": 0.0, "PENDING": 0.0}
    for b in bills:
        amt = b["grand_total"]
        if b["payment_status"] == "PAID":
            st["PAID"] += amt
        elif b["payment_status"] == "PARTIALLY_PAID":
            st["PARTIAL"] += amt
        elif b["payment_status"] != "CANCELLED":
            st["PENDING"] += amt
    payment_status = [
        {"name": "Paid", "value": round(st["PAID"], 2), "color": "#16A34A"},
        {"name": "Partial", "value": round(st["PARTIAL"], 2), "color": "#F59E0B"},
        {"name": "Unpaid", "value": round(st["PENDING"], 2), "color": "#DC2626"},
    ]
    payment_status = [s for s in payment_status if s["value"] > 0]

    pending_gps = []
    for gp in gate_passes:
        dm = del_map.get(gp["id"], {})
        pending_detail = []
        total_pending = 0
        for it in gp["items"]:
            delivered = dm.get(it["item_name"], 0)
            p = max(0, it["received_qty"] - delivered)
            if p > 0:
                total_pending += p
                pending_detail.append({
                    "item_name": it["item_name"],
                    "specification": it.get("specification") or "",
                    "category": it.get("category") or "",
                    "received": it["received_qty"],
                    "delivered": delivered,
                    "pending": p,
                })
        if total_pending > 0:
            pending_gps.append(
                {
                    "gate_pass_number": gp["gate_pass_number"],
                    "client_name": gp["client_name"],
                    "pending": total_pending,
                    "items": pending_detail,
                }
            )
    pending_gps.sort(key=lambda x: -x["pending"])
    pending_gps = pending_gps[:6]

    return {
        "revenue": round(revenue, 2),
        "collected": round(collected, 2),
        "outstanding": round(outstanding, 2),
        "collectionRate": round(collection_rate, 2),
        "avgBill": round(avg, 2),
        "billCount": bill_count,
        "paidBills": paid,
        "partialBills": partial,
        "pendingBills": pending,
        "gatePasses": gp_count,
        "itemsReceived": items_received,
        "itemsDelivered": items_delivered,
        "itemsPending": items_pending,
        "activeClients": active_clients,
        "topClients": top_clients,
        "revenueSeries": revenue_series,
        "paymentStatus": payment_status,
        "pendingGatePasses": pending_gps,
    }


@router.get("/dashboard/summary")
async def get_dashboard_summary(
    period: str = Query("month", pattern="^(day|week|month|quarter|year)$"),
    current_user: dict = Depends(require_capability("dashboard:read")),
):
    cur_start, cur_end, prev_start, prev_end = _windows(period)

    cur_bills = await _fetch_bills(cur_start, cur_end)
    prev_bills = await _fetch_bills(prev_start, prev_end)
    cur_gps = await _fetch_gate_passes(cur_start, cur_end)
    cur_dels = await _fetch_deliveries([gp["id"] for gp in cur_gps])
    prev_gps = await _fetch_gate_passes(prev_start, prev_end)
    prev_dels = await _fetch_deliveries([gp["id"] for gp in prev_gps])

    current = aggregate(cur_bills, cur_gps, cur_dels, period)
    previous = aggregate(prev_bills, prev_gps, prev_dels, period)

    cur_clients = {b["client_name"] for b in cur_bills}
    prev_clients = {b["client_name"] for b in prev_bills}
    current["newClients"] = len(cur_clients - prev_clients)
    previous["newClients"] = 0

    for m in (current, previous):
        m["quotationsDraft"] = 0
        m["quotationsSent"] = 0
        m["quotationsAccepted"] = 0
        m["quotationAcceptedValue"] = 0.0

    return {"current": current, "previous": previous}


# ── Recent Activity ───────────────────────────────────────────────────────────

@router.get("/dashboard/recent-activity")
async def recent_activity(
    limit: int = Query(default=10, ge=1, le=50),
    current_user: dict = Depends(require_capability("dashboard:read")),
):
    cursor = audit_collection.find().sort("timestamp", -1).limit(limit)
    activities = []
    async for doc in cursor:
        ts = doc.get("timestamp")
        activities.append({
            "id": str(doc.get("_id", "")),
            "user_id": doc.get("user_id") or doc.get("auth_id") or "system",
            "action": doc.get("action") or "UNKNOWN",
            "entity": doc.get("entity_type") or doc.get("entity") or "",
            "entity_id": str(doc.get("entity_id", "")),
            "details": doc.get("details") or {},
            "timestamp": ts.isoformat() if ts else None,
        })
    return activities


# ── Outstanding Aging ─────────────────────────────────────────────────────────

@router.get("/dashboard/outstanding-aging")
async def outstanding_aging(
    current_user: dict = Depends(require_capability("dashboard:read")),
):
    from datetime import timedelta
    now = datetime.now(timezone.utc)
    buckets = {"current": 0.0, "30_day": 0.0, "60_day": 0.0, "90_day": 0.0, "over_90": 0.0}

    cursor = bills_collection.find(
        {"payment_status": {"$in": ["PENDING", "PARTIALLY_PAID", "ISSUED"]}},
        {"outstanding_amount": 1, "created_at": 1, "_id": 0},
    )
    async for doc in cursor:
        outstanding = float(doc.get("outstanding_amount") or 0)
        if outstanding <= 0:
            continue
        created = doc.get("created_at")
        if not created:
            buckets["current"] += outstanding
            continue
        created_utc = _utc(created)
        if not created_utc:
            buckets["current"] += outstanding
            continue
        age_days = (now - created_utc).days
        if age_days <= 0:
            buckets["current"] += outstanding
        elif age_days <= 30:
            buckets["30_day"] += outstanding
        elif age_days <= 60:
            buckets["60_day"] += outstanding
        elif age_days <= 90:
            buckets["90_day"] += outstanding
        else:
            buckets["over_90"] += outstanding

    return {
        "current": round(buckets["current"], 2),
        "30_day": round(buckets["30_day"], 2),
        "60_day": round(buckets["60_day"], 2),
        "90_day": round(buckets["90_day"], 2),
        "over_90": round(buckets["over_90"], 2),
    }


# ── 12-Month Trend ───────────────────────────────────────────────────────────

@router.get("/dashboard/yearly-trend")
async def yearly_trend(
    current_user: dict = Depends(require_capability("dashboard:read")),
):
    from datetime import timedelta
    now = datetime.now(timezone.utc)
    months = []
    for i in range(11, -1, -1):
        d = now - timedelta(days=30 * i)
        months.append({
            "key": d.strftime("%Y-%m"),
            "label": d.strftime("%b %y"),
            "revenue": 0.0,
            "collected": 0.0,
            "bills": 0,
        })

    month_index = {m["key"]: m for m in months}

    cursor = bills_collection.find(
        {"payment_status": {"$ne": "CANCELLED"}},
        {"grand_total": 1, "paid_amount": 1, "created_at": 1, "_id": 0},
    )
    async for doc in cursor:
        created = doc.get("created_at")
        if not created:
            continue
        created_utc = _utc(created)
        if not created_utc:
            continue
        key = created_utc.strftime("%Y-%m")
        if key in month_index:
            month_index[key]["revenue"] += float(doc.get("grand_total") or 0)
            month_index[key]["collected"] += float(doc.get("paid_amount") or 0)
            month_index[key]["bills"] += 1

    for m in months:
        m["revenue"] = round(m["revenue"], 2)
        m["collected"] = round(m["collected"], 2)

    return months


# ── Today's Deliveries ────────────────────────────────────────────────────────

@router.get("/dashboard/today-deliveries")
async def today_deliveries(
    current_user: dict = Depends(require_capability("dashboard:read")),
):
    """Today's delivery breakdown by client with current balance."""
    from datetime import timedelta as td

    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + td(days=1)

    # Aggregate today's deliveries by client
    client_map: Dict[str, dict] = {}

    del_cursor = deliveries_collection.find({
        "created_at": {"$gte": today_start, "$lt": today_end},
        "status": {"$ne": "CANCELLED"},
    })
    async for doc in del_cursor:
        try:
            dl = _decrypt_del(doc)
            client = dl.get("client_name", "").strip()
            if not client:
                continue
            if client not in client_map:
                client_map[client] = {
                    "client_name": client,
                    "items": {},
                    "total_qty": 0,
                    "gate_pass_ids": set(),
                }
            for item in dl.get("items", []):
                item_name = item.get("item_name", "")
                spec = item.get("specification") or ""
                qty = item.get("quantity", 0)
                detail_key = f"{item_name}||{spec}" if spec else item_name
                if detail_key not in client_map[client]["items"]:
                    client_map[client]["items"][detail_key] = {
                        "item_name": item_name,
                        "specification": spec,
                        "quantity": 0,
                    }
                client_map[client]["items"][detail_key]["quantity"] += qty
                client_map[client]["total_qty"] += qty
            gp_id = dl.get("gate_pass_id")
            if gp_id:
                client_map[client]["gate_pass_ids"].add(gp_id)
        except Exception:
            continue

    # Get outstanding balance per client
    bill_cursor = bills_collection.find({"payment_status": {"$ne": "CANCELLED"}})
    async for doc in bill_cursor:
        try:
            bill = _decrypt_bill(doc)
            client = bill.get("client_name", "").strip()
            if client in client_map:
                client_map[client]["outstanding"] = client_map[client].get("outstanding", 0) + (bill.get("outstanding_amount", 0) or 0)
        except Exception:
            continue

    # Get pending items per client (from all open gate passes)
    gp_cursor = gatepasses_collection.find({"status": {"$ne": "CANCELLED"}})
    async for doc in gp_cursor:
        try:
            gp = _decrypt_gp(doc)
            client = gp.get("client_name", "").strip()
            if client not in client_map:
                continue
            for item in gp.get("items", []):
                item_name = item.get("item_name", "")
                spec = item.get("specification") or ""
                received = item.get("received_qty", 0)
                detail_key = f"{item_name}||{spec}" if spec else item_name
                if detail_key not in client_map[client].get("pending_items", {}):
                    client_map[client].setdefault("pending_items", {})[detail_key] = {
                        "item_name": item_name,
                        "specification": spec,
                        "received": 0,
                        "delivered": 0,
                    }
                client_map[client]["pending_items"][detail_key]["received"] += received
        except Exception:
            continue

    # Subtract delivered quantities from pending
    del_cursor2 = deliveries_collection.find({"status": {"$ne": "CANCELLED"}})
    async for doc in del_cursor2:
        try:
            dl = _decrypt_del(doc)
            client = dl.get("client_name", "").strip()
            if client not in client_map:
                continue
            for item in dl.get("items", []):
                item_name = item.get("item_name", "")
                spec = item.get("specification") or ""
                qty = item.get("quantity", 0)
                detail_key = f"{item_name}||{spec}" if spec else item_name
                if detail_key in client_map.get(client, {}).get("pending_items", {}):
                    client_map[client]["pending_items"][detail_key]["delivered"] += qty
        except Exception:
            continue

    # Build result
    results = []
    for client, data in client_map.items():
        items_list = list(data["items"].values())
        pending_items = []
        for pi in data.get("pending_items", {}).values():
            pi["pending"] = max(0, pi["received"] - pi["delivered"])
            if pi["pending"] > 0:
                pending_items.append(pi)

        results.append({
            "client_name": client,
            "delivered_items": items_list,
            "total_qty": data["total_qty"],
            "gate_pass_count": len(data.get("gate_pass_ids", set())),
            "outstanding": round(data.get("outstanding", 0), 2),
            "pending_items": pending_items,
        })

    results.sort(key=lambda x: x["total_qty"], reverse=True)
    return results


# ── CSV Export ────────────────────────────────────────────────────────────────


@router.get("/export/gatepasses")
async def export_gatepasses_csv(
    client_name: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(require_capability("report:read")),
):
    """Export gate passes as CSV for download."""
    query: dict = {}
    if client_name:
        query["client_name_search"] = get_search_token(client_name)
    if status:
        query["status"] = status

    cursor = gatepasses_collection.find(query).sort("created_at", -1)
    rows = []
    async for doc in cursor:
        dec = _decrypt_gp(doc)
        for item in dec.get("items", []):
            rows.append({
                "gate_pass_number": dec.get("gate_pass_number", ""),
                "client_name": dec.get("client_name", ""),
                "receiving_date": str(dec.get("receiving_date", ""))[:10],
                "status": dec.get("status", ""),
                "item_name": item.get("item_name", ""),
                "specification": item.get("specification", ""),
                "category": item.get("category", ""),
                "client_qty": item.get("client_qty", 0),
                "received_qty": item.get("received_qty", 0),
                "difference": item.get("difference", 0),
            })

    if not rows:
        rows.append({"gate_pass_number": "No data", "client_name": "", "receiving_date": "",
                       "status": "", "item_name": "", "specification": "", "category": "",
                       "client_qty": 0, "received_qty": 0, "difference": 0})

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=rows[0].keys())
    writer.writeheader()
    writer.writerows(rows)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=gate-passes.csv"},
    )


@router.get("/export/bills")
async def export_bills_csv(
    client_name: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(require_capability("report:read")),
):
    """Export bills as CSV for download."""
    query: dict = {}
    if client_name:
        query["client_name_search"] = get_search_token(client_name)
    if status:
        query["payment_status"] = status

    cursor = bills_collection.find(query).sort("created_at", -1)
    rows = []
    async for doc in cursor:
        try:
            dec = decrypt_dict(doc, SENSITIVE_FIELDS_BILL)
        except Exception:
            continue
        dec["id"] = str(dec["_id"])
        del dec["_id"]
        for item in dec.get("items", []):
            rows.append({
                "bill_number": dec.get("bill_number", ""),
                "client_name": dec.get("client_name", ""),
                "bill_date": str(dec.get("bill_date", ""))[:10],
                "status": dec.get("payment_status", ""),
                "item_name": item.get("item_name", ""),
                "quantity": item.get("quantity", 0),
                "unit_price": item.get("unit_price", 0),
                "total": item.get("total", 0),
                "grand_total": dec.get("grand_total", 0),
                "amount_paid": dec.get("amount_paid", 0),
                "balance": dec.get("balance", 0),
            })

    if not rows:
        rows.append({"bill_number": "No data", "client_name": "", "bill_date": "",
                       "status": "", "item_name": "", "quantity": 0, "unit_price": 0,
                       "total": 0, "grand_total": 0, "amount_paid": 0, "balance": 0})

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=rows[0].keys())
    writer.writeheader()
    writer.writerows(rows)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=bills.csv"},
    )


@router.get("/export/deliveries")
async def export_deliveries_csv(
    client_name: Optional[str] = None,
    current_user: dict = Depends(require_capability("report:read")),
):
    """Export deliveries as CSV for download."""
    query: dict = {}
    if client_name:
        query["client_name_search"] = get_search_token(client_name)

    cursor = deliveries_collection.find(query).sort("created_at", -1)
    rows = []
    async for doc in cursor:
        try:
            dec = decrypt_dict(doc, SENSITIVE_FIELDS_DEL)
        except Exception:
            continue
        dec["id"] = str(dec["_id"])
        del dec["_id"]
        for item in dec.get("items", []):
            rows.append({
                "client_name": dec.get("client_name", ""),
                "delivery_date": str(dec.get("delivery_date", ""))[:10],
                "delivered_by": dec.get("delivered_by", ""),
                "received_by": dec.get("received_by", ""),
                "status": dec.get("status", ""),
                "item_name": item.get("item_name", ""),
                "specification": item.get("specification", ""),
                "quantity": item.get("quantity", 0),
            })

    if not rows:
        rows.append({"client_name": "No data", "delivery_date": "", "delivered_by": "",
                       "received_by": "", "status": "", "item_name": "", "specification": "",
                       "quantity": 0})

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=rows[0].keys())
    writer.writeheader()
    writer.writerows(rows)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=deliveries.csv"},
    )


# ── Excel Export ──────────────────────────────────────────────────────────────

HEADER_FONT = None
HEADER_FILL = None
HEADER_ALIGN = None
THIN_BORDER = None


def _init_openpyxl():
    global HEADER_FONT, HEADER_FILL, HEADER_ALIGN, THIN_BORDER
    if HEADER_FONT is not None:
        return
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def _style_sheet(ws, headers, rows):
    _init_openpyxl()
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
    for col_idx, h in enumerate(headers, 1):
        max_len = max(len(str(h)), *(len(str(r[col_idx - 1])) for r in rows) if rows else [0])
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)
    ws.auto_filter.ref = ws.dimensions


@router.get("/export/gatepasses/xlsx")
async def export_gatepasses_xlsx(
    client_name: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(require_capability("report:read")),
):
    """Export gate passes as Excel."""
    from openpyxl import Workbook

    query: dict = {}
    if client_name:
        query["client_name_search"] = get_search_token(client_name)
    if status:
        query["status"] = status

    cursor = gatepasses_collection.find(query).sort("created_at", -1)
    rows = []
    async for doc in cursor:
        dec = _decrypt_gp(doc)
        for item in dec.get("items", []):
            rows.append([
                dec.get("gate_pass_number", ""),
                dec.get("client_name", ""),
                str(dec.get("receiving_date", ""))[:10],
                dec.get("status", ""),
                item.get("item_name", ""),
                item.get("specification", ""),
                item.get("category", ""),
                item.get("client_qty", 0),
                item.get("received_qty", 0),
                item.get("difference", 0),
            ])

    wb = Workbook()
    ws = wb.active
    ws.title = "Gate Passes"
    headers = ["Gate Pass #", "Client", "Date", "Status", "Item", "Specification", "Category", "Client Qty", "Received", "Difference"]
    if rows:
        _style_sheet(ws, headers, rows)
    else:
        ws.cell(row=1, column=1, value="No data")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=gate-passes.xlsx"},
    )


@router.get("/export/bills/xlsx")
async def export_bills_xlsx(
    client_name: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(require_capability("report:read")),
):
    """Export bills as Excel."""
    from openpyxl import Workbook

    query: dict = {}
    if client_name:
        query["client_name_search"] = get_search_token(client_name)
    if status:
        query["payment_status"] = status

    cursor = bills_collection.find(query).sort("created_at", -1)
    rows = []
    async for doc in cursor:
        try:
            dec = decrypt_dict(doc, SENSITIVE_FIELDS_BILL)
        except Exception:
            continue
        for item in dec.get("items", []):
            rows.append([
                dec.get("bill_number", ""),
                dec.get("client_name", ""),
                str(dec.get("bill_date", ""))[:10],
                dec.get("payment_status", ""),
                item.get("item_name", ""),
                item.get("quantity", 0),
                item.get("unit_price", 0),
                item.get("total", 0),
                dec.get("grand_total", 0),
                dec.get("amount_paid", 0),
                dec.get("balance", 0),
            ])

    wb = Workbook()
    ws = wb.active
    ws.title = "Bills"
    headers = ["Bill #", "Client", "Date", "Status", "Item", "Qty", "Unit Price", "Item Total", "Grand Total", "Paid", "Balance"]
    if rows:
        _style_sheet(ws, headers, rows)
    else:
        ws.cell(row=1, column=1, value="No data")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bills.xlsx"},
    )


@router.get("/export/deliveries/xlsx")
async def export_deliveries_xlsx(
    client_name: Optional[str] = None,
    current_user: dict = Depends(require_capability("report:read")),
):
    """Export deliveries as Excel."""
    from openpyxl import Workbook

    query: dict = {}
    if client_name:
        query["client_name_search"] = get_search_token(client_name)

    cursor = deliveries_collection.find(query).sort("created_at", -1)
    rows = []
    async for doc in cursor:
        try:
            dec = decrypt_dict(doc, SENSITIVE_FIELDS_DEL)
        except Exception:
            continue
        for item in dec.get("items", []):
            rows.append([
                dec.get("client_name", ""),
                str(dec.get("delivery_date", ""))[:10],
                dec.get("delivered_by", ""),
                dec.get("received_by", ""),
                dec.get("status", ""),
                item.get("item_name", ""),
                item.get("specification", ""),
                item.get("quantity", 0),
            ])

    wb = Workbook()
    ws = wb.active
    ws.title = "Deliveries"
    headers = ["Client", "Date", "Delivered By", "Received By", "Status", "Item", "Specification", "Qty"]
    if rows:
        _style_sheet(ws, headers, rows)
    else:
        ws.cell(row=1, column=1, value="No data")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=deliveries.xlsx"},
    )
